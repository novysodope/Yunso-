import requests
from datetime import datetime, timedelta
from wxauto import *
import os
import json
from openai import OpenAI
import re
from rich import print
import time
from rich.console import Console
from datetime import datetime
import sys
from rich.progress import Progress
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from bs4 import BeautifulSoup
import configparser
import threading

config_file = 'yunso_config.ini'
config = configparser.ConfigParser()
config.read(config_file, encoding="utf-8")


class Yunso:
    def __init__(self,config):
        global config_file
        self.banner = """
____    ____  __    __  .__   __.      _______.  ______   
\\   \\  /   / |  |  |  | |  \\ |  |     /       | /  __  \\  
 \\   \\/   /  |  |  |  | |   \\|  |    |   (----`|  |  |  | 
  \\_    _/   |  |  |  | |  . `  |     \\   \\    |  |  |  | 
    |  |     |  `--'  | |  |\\   | .----)   |   |  `--'  | 
    |__|      \\______/  |__| \\__| |_______/     \\______/  

https://github.com/novysodope/yunso      云索|Fupo Series

"""
        print(f"[green]{self.banner}[/green]")
        print("————————————————初始化开始————————————————\n")
        print(f"[green][配置文件][/green] {config_file}\n")
#---------------------配置-------------------------------
        self.config = config
        self.listen_list = self.config.get("DEFAULT", "listen_list", fallback="").split(",")
        item = '|'.join(self.listen_list)
        print(f"[green][消息发送对象][/green] {item}\n")
        self.access_token = self.config['DEFAULT']['github_api_key']
        print(f"[green][Github访问令牌][/green] {self.access_token}\n")

        self.SEARCH_TYPES = ['repositories']
        self.GITHUB_API_URL = 'https://api.github.com/search'
        self.VULBOX_API_URL = "https://vip.vulbox.com/api/data/base_vuln_list"

        self.SENT_RECORDS_FILE = "sent_records.json"
        self.LAST_REPORT_FILE = "last_report.json"

        self.now = datetime.now()
        self.formatted_date = self.now.strftime("%Y-%m-%d")
        self.filename = f"推送周报_{self.formatted_date}.xlsx"
        self.SEARCH_KEYWORDS = self.config.get("DEFAULT", "search_keywords", fallback="").split(",")
        itemkey = '|'.join(self.SEARCH_KEYWORDS)
        print(f"[green][监控内容][/green] {itemkey}\n")
        self.weekdays = {
            0: "星期一",
            1: "星期二",
            2: "星期三",
            3: "星期四",
            4: "星期五",
            5: "星期六",
            6: "星期日"
        }
        self.last_printed_date = None

        self.GITHUB_HEADERS = {
            'Authorization': f'token {self.access_token}',
            'Accept': 'application/vnd.github.v3+json',
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36"
        }

        self.VULBOX_HEADERS = {
            "Accept": "application/json, text/plain, */*",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.131 Safari/537.36",
            "Content-Type": "application/json; charset=UTF-8",
            "Origin": "https://vip.vulbox.com",
            "Sec-Fetch-Site": "same-origin",
            "Sec-Fetch-Mode": "cors",
            "Sec-Fetch-Dest": "empty",
            "Referer": "https://vip.vulbox.com/originalList?keyword=",
            "Accept-Language": "zh-CN,zh;q=0.9"
        }
        self.wx = WeChat()
        print('Yunso start')
#---------------------配置-------------------------------
     
#----------------------初始化sent_records.json，用来检查是否有重复记录------------------------------
        if os.path.exists(self.SENT_RECORDS_FILE):
            try:
                with open(self.SENT_RECORDS_FILE, "r") as f:
                    content = f.read().strip()
                    self.sent_records = set(json.loads(content)) if content else set()
            except json.JSONDecodeError:
                self.sent_records = set()
        else:
            self.sent_records = set()
#----------------------初始化sent_records.json，用来检查是否有重复记录------------------------------

#----------------------拿到漏洞标题------------------------------
    def query_vulbox(self, cve):
        api_url = self.VULBOX_API_URL
        payload = {
            "keyword": cve,
            "riskLevelSort": "",
            "modifyTimeSort": 0,
            "publishTimeSort": None,
            "page": 1,
            "pageSize": 10
        }
        try:
            response = requests.post(api_url, json=payload, headers=self.VULBOX_HEADERS, timeout=10)
            time.sleep(3)

            if response.status_code == 200:
                data = response.json()
                records = data.get("data", {}).get("records", [])

                for record in records:
                    print(f"input:{cve}")
                    vuln_cve = record.get("vulnCveCode", "")
                    print(f"api:{vuln_cve}")

                    if vuln_cve == cve:
                        vuln_id = records[0].get("id", "")
                        print("|--------Block begin---------|")
                        print(f"input:{cve}")
                        print(f"api:{vuln_cve}")
                        print(f"id:{vuln_id}")

                        if vuln_id:
                            detail_url = f"https://vip.vulbox.com/detail/{vuln_id}"
                            print(f"url:{detail_url}")
                            detail_response = requests.get(detail_url, headers=self.VULBOX_HEADERS)

                            if detail_response.status_code == 200:
                                soup = BeautifulSoup(detail_response.text, 'html.parser')
                                title_tag = soup.find('h3', class_='break-all ui-pr60 header-tit text-overflow-3')

                                if title_tag and title_tag.has_attr('title'):
                                    vuln_name = title_tag['title']
                                    print(f"name:{vuln_name}")
                                    print("|------------End-------------|\n")
                                    return vuln_name
                    else:
                        return cve
        except Exception as e:
            print(f"查询接口出错: {e}")
        return cve
#----------------------拿到漏洞标题------------------------------

#-----------------------Github监听，结果整理-----------------------------
    def search_and_notify(self):
        global sent_records
        new_sent_records = set()
        message_vuln = []
        nows = datetime.utcnow()
        search_time = nows - timedelta(days=2)
        search_time_td = search_time.strftime('%Y-%m-%d')

        with Progress() as progress:
            search_keywords_count = len(self.SEARCH_KEYWORDS)
            task = progress.add_task(f"[cyan]一共[{search_keywords_count}]个关键字，请稍候[/cyan]", total=len(self.SEARCH_KEYWORDS))
            sys.stdout.write("\r")
            sys.stdout.flush()
            for keyword in self.SEARCH_KEYWORDS:
                query = f'{keyword.lower()} created:>{search_time_td}'
                params = {'q': query}

                for search_type in self.SEARCH_TYPES:
                    full_url = f"{self.GITHUB_API_URL}/{search_type}"
                    response = requests.get(full_url, headers=self.GITHUB_HEADERS, params=params, timeout=120)
                    time.sleep(5)

                    if response.status_code == 200:
                        results = response.json()
                        items = results.get('items', [])

                        for item in items:
                            name = item.get('name', '')
                            html_url = item.get('html_url', '')
                            description = item.get('description', '')
                            unique_identifier = f"{name}|{html_url}|{description}"

                            if unique_identifier not in self.sent_records and unique_identifier not in new_sent_records:
                                if 'cve' in name.lower():
                                    match = re.search(r'cve[-_ ]?\d{4}[-_ ]?\d{3,}', name, re.IGNORECASE)
                                    if match:
                                        cve = match.group(0).upper().replace("_", "-").replace(" ", "-")
                                        cve_name = self.query_vulbox(cve)
                                        if cve_name:
                                            name = cve_name
                                            message_vuln.append(
                                                f"- 漏洞信息: {name}\n- 地址: {html_url}\n- 描述: {description}\n")
                                        else:
                                            message_vuln.append(
                                                f"- 名称: {name}\n- 地址: {html_url}\n- 描述: {description}\n")
                                else:
                                    message_vuln.append(f"- 名称: {name}\n- 地址: {html_url}\n- 描述: {description}\n")
                                new_sent_records.add(unique_identifier)

                    else:
                        print(f'请求失败，状态码: {response.status_code}, 响应: {response.text}')

                progress.update(task, advance=1)  # 更新进度条
 #-----------------------Github监听，结果整理-----------------------------

 #-----------------------发送文本-----------------------------
        if message_vuln:
            msg = '#新动态提示\n\n' + '\n'.join(message_vuln)
            msg += '\n* POC及工具等信息均来自Github，请注意辨别。本脚本仅用于监控推送'
            for who in self.listen_list:
                print(f"正在向 {who} 发送消息...")
                self.wx.SendMsg(msg, who)
            self.sent_records.update(new_sent_records)

            with open(self.SENT_RECORDS_FILE, "w", encoding="utf-8") as f:
                json.dump(list(self.sent_records), f)
        else:
            print("没有新内容，开始表明存活状态")
            self.wx.GetSessionList()
            self.wx.SendMsg("online", '文件传输助手')
#-----------------------发送文本-----------------------------

#---------------------生成周报-------------------------------
    def parse_sent_records(self):
        print(self.SENT_RECORDS_FILE)
        if not os.path.exists(self.SENT_RECORDS_FILE):
            return []

        with open(self.SENT_RECORDS_FILE, "r") as f:
            try:
                records = json.load(f)
            except json.JSONDecodeError:
                records = []
        parsed_records = []

        for record in records:
            parts = record.split('|')
            if len(parts) == 3:
                title, address, description = parts
                parsed_records.append({"标题": title, "地址": address, "描述": description})

        return parsed_records

    def get_previous_report(self):
        if os.path.exists(self.LAST_REPORT_FILE):
            try:
                with open(self.LAST_REPORT_FILE, "r", encoding="utf-8") as f:
                    return set(tuple(item.items()) for item in json.load(f))
            except json.JSONDecodeError:
                return set()
        return set()

    def filter_new_records(self, records):
        previous_records = self.get_previous_report()
        new_records = [record for record in records if tuple(record.items()) not in previous_records]
        return new_records

    def generate_weekly_report(self):
        records = self.parse_sent_records()
        if not records:
            print("没有新漏洞记录，无需生成周报。")
            return False

        new_records = self.filter_new_records(records)

        if not new_records:
            print("本周无新增漏洞记录，跳过生成。")
            return False

        df = pd.DataFrame(records)
        df.to_excel(self.filename, index=False)

        wb = load_workbook(self.filename)
        ws = wb.active
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = max(adjusted_width, 54)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)
                cell.border = Border(
                    top=Side(border_style="thin", color="000000"),
                    bottom=Side(border_style="thin", color="000000"),
                    left=Side(border_style="thin", color="000000"),
                    right=Side(border_style="thin", color="000000")
                    )

        wb.save(self.filename)

        print(f"周报已生成：{self.filename}")

        with open(self.LAST_REPORT_FILE, "w", encoding="utf-8") as f:
            json.dump(records, f, ensure_ascii=False, indent=4)

        msg = '开始统计本周周报：'
        for who in self.listen_list:
            print(f"正在向 {who} 发送消息...")
            wxchat.SendMsg(msg, who)
            print("消息发送完成。")
        return True
#---------------------生成周报-------------------------------

#-----------------------每周五发送文件-----------------------------
    def print_weekday(self):
        print("周五开始统计周报")
        if self.generate_weekly_report():
            wxchat.SendFiles(self.filename, self.who)
#-----------------------每周五发送文件-----------------------------

#----------------------脚本循环------------------------------
    def run(self):
        while True:
            self.search_and_notify()
            if self.now.weekday() == 4 and self.last_printed_date != self.now.date():
                self.print_weekday()
                self.last_printed_date = self.now.date()
            print("等待30分钟后继续...")
            time.sleep(1800)
#----------------------脚本循环------------------------------

class Yunsobt:
#---------------------配置-------------------------------
    GITHUB_API_URL = "https://api.github.com/search/repositories?q={query}&sort=stars&order=desc"
    def __init__(self,config):
        self.config = config
        self.access_token = self.config['DEFAULT']['github_api_key']
        self.listen_list = self.config.get("DEFAULT", "listen_list", fallback="").split(",")
        self.name = self.config['DEFAULT']['me']
        self.me = '@' + self.name
        print(f"[green][机器人名称][/green] {self.me}\n")
        self.replied_msgs = set()
        self.black_keywords = self.config.get("DEFAULT", "black_keywords", fallback="").split(",")
        itembalck = '|'.join(self.black_keywords)
        print(f"[green][黑名单关键字][/green] {itembalck}\n")
        print('Yunsobt start')
#---------------------配置-------------------------------
        self.wx = WeChat()
        for i in self.listen_list:
            self.wx.AddListenChat(who=i)

    def run(self):
        while True:
#---------------------监听消息-------------------------------
            try:
                msgs = self.wx.GetListenMessage()
            except Exception as e:
                for i in self.listen_list:
                    self.wx.AddListenChat(who=i)
                    msgs = self.wx.GetListenMessage()
            for chat in msgs:
                who = chat.who
                one_megs = msgs.get(chat)
                for msg in one_megs:
                    content = msg.content
#---------------------监听消息-------------------------------
                    # 因为这里是按秒来循环监听的，所以这里要加个标识符，防止特定消息被反复处理
                    msg_id = msg.id

                    if msg_id in self.replied_msgs:
                        continue
#---------------------处理收到的关键字-------------------------------
                    if self.me in content:
                        chat.SendMsg("\n请稍后，正在查询...", who)
                        time.sleep(1)
                        query = content.replace(self.me, '').strip()
                        if query:
                            headers = {"Authorization": f"token {self.access_token}"}
                            response = requests.get(self.GITHUB_API_URL.format(query=query), headers=headers, timeout=120)
                            if response.status_code == 200:
                                result = response.json()
                                if "items" in result and result["items"]:
                                    filtered_results = []
                                    for item in result["items"][:5]:
                                        name = item['name']
                                        url = item['html_url']
                                        description = item.get('description', '无描述')
                                        try:
                                            # 这里还是有点问题，有时间再改，先加个try
                                            if any(keyword.lower() in (name + description).lower() for keyword in self.black_keywords):
                                                continue
                                        except Exception as e:
                                            print(e)
                                            pass

                                        filtered_results.append(f"- 标题: {name}\n- URL: {url}\n- 描述: {description}\n\n")

                                    if filtered_results:
                                        chat.SendMsg(f"\n只显示star前5:\n\n{''.join(filtered_results)}", who)
                                    else:
                                        chat.SendMsg("搜索结果中无合适内容", who)
                                else:
                                    chat.SendMsg("没有找到相关仓库", who)
                            else:
                                chat.SendMsg("GitHub 搜索失败", who)

                        self.replied_msgs.add(msg_id)

            time.sleep(1)
#---------------------处理收到的关键字-------------------------------

#---------------------AI-------------------------------
class AiBot:
    def __init__(self,config):
        self.config = config
        self.ai_name = self.config['DEFAULT']['me']
        self.ai_key = self.config['AI']['ai_key']
        print(f"[green][AI KEY][/green] {self.ai_key}\n")
        self.ai_url = self.config['AI']['ai_url']
        print(f"[green][AI API][/green] {self.ai_url}\n")
        self.listen_list = self.config.get("DEFAULT", "listen_list", fallback="").split(",")
        self.replied_msgs = set()
        self.wx = WeChat()
        self.messages = []
        print('AiBot start')
        print("————————————————初始化结束————————————————\n")
        for i in self.listen_list:
            self.wx.AddListenChat(who=i)

    def run(self):
        while True:
            try:
                msgs = self.wx.GetListenMessage()
            except Exception as e:
                for i in self.listen_list:
                    self.wx.AddListenChat(who=i)
                    msgs = self.wx.GetListenMessage()
            for chat in msgs:
                who = chat.who
                one_megs = msgs.get(chat)
                for msg in one_megs:
                    content = msg.content
                    #print(content)
                    msg_id = msg.id
                    if msg_id in self.replied_msgs:
                        continue
                    what = '二狗， '
                    stra = '二狗'
                    if content == stra:
                        print(content)
                        self.wx.SendMsg(
                            '什么事？如果想使用AI，请在问题前面加上“二狗”+中文逗号+空格，比如：二狗[，] 在吗？如果想用新会话，请发“狗蛋[，]清除”',
                            who)
                    if what in content:
                        print(content)
                        #query = content.replace(what, '').strip()
                        self.wx.SendMsg(
                            '思考中，请稍后，如果超过1分钟再重新询问（AI思考时间有点长加上回答内容长，所以需要时间）',
                            who)
                        try:
                            self.messages.append({'role': 'user', 'content': content})

                            client = OpenAI(api_key=self.ai_key,
                                            base_url=self.ai_url)
                            response = client.chat.completions.create(
                                model="deepseek-v3",
                                messages=self.messages
                            )
                            print(response.choices[0].message.content)
                            self.wx.SendMsg(response.choices[0].message.content, who)

                            self.messages.append({'role': 'assistant', 'content': response.choices[0].message.content})
                        except Exception as e:
                            print(e)
                            pass

                    if '狗蛋，清除' in content:
                        self.messages = []
                        print(self.messages)
                        self.wx.SendMsg("上下文信息清除完毕，聊天重置", who)


#---------------------AI-------------------------------


if __name__ == "__main__":
    yunso = Yunso(config)
    yunsobt = Yunsobt(config)
    aibot = AiBot(config)
    # 又学会了一招，设置daemon=True让线程随主线程退出
    yunso_thread = threading.Thread(target=yunso.run,daemon=True)
    yunsobt_thread = threading.Thread(target=yunsobt.run,daemon=True)
    aibot_thread = threading.Thread(target=aibot.run, daemon=True)

    yunso_thread.start()

    yunsobt_thread.start()

    aibot_thread.start()

    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        exit = '''
        \n\n
        [red]ヾ(￣▽￣)Bye~Bye~[/red]
        \n\n
    '''
        print(exit)
